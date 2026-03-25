"""
uploader_qoo10.py – Qoo10 J·QSM 대량등록용 엑셀 생성 & Drive 업로드
Project: Plan B Cabinet – Qoo10 Japan Beauty Sourcing
Version: 0.9.4

변경사항 (v0.9.3 → v0.9.4):
  - [FIX] Q열: 로컬 경로(artifacts/) → 원본 thumbnail URL 폴백
  - [FIX] V/W열: header_html, footer_html 2,500자 제한 + Python 코드 혼입 제거
  - [FIX] Y열: 배송비 코드 → 환경변수에서 숫자 코드 읽기 (기본값 813137)
  - [FIX] AJ열: 무게 g→kg 변환 (소수점 둘째자리, 최대 30kg)
  - [FIX] bytes 타입 방어 (thumbnail_processed bytes 이슈)

기존 유지 (v0.9.3):
  - seller_id → item.get('item_id') 수정
  - 변수명 통일 (qoo10_cat, qoo10_brand, sell_price 등)
  - generate_qoo10_excel 함수 통합 (등록코드 제외 로직 포함)
  - 약기법 필터, 50컬럼, $$ 구분자, 상품명 50자 제한
"""

import json
import math
import logging
import os
import re
from io import BytesIO
from datetime import datetime

logger = logging.getLogger(__name__)

# ★ 약기법 필터 import
try:
    from yakujiho_filter import sanitize_jp
    YAKUJIHO_ENABLED = True
    logger.info("[약기법] yakujiho_filter 로드 완료 (uploader)")
except ImportError:
    logger.warning("[약기법] yakujiho_filter 모듈 없음 – 필터링 비활성화")
    YAKUJIHO_ENABLED = False
    def sanitize_jp(text):
        return text, [], 0


# ── 상수 ──────────────────────────────────────────────────────
# ★ v0.9.4: 배송비 코드 → 환경변수에서 숫자 코드 읽기
QOO10_KSE_SHIPPING_CODE = os.environ.get("QOO10_KSE_SHIPPING_CODE", "813137")

# ★ v0.9.4: V/W열 글자 제한
HEADER_MAX_LEN = 2500
FOOTER_MAX_LEN = 2500

# 공식 50컬럼 헤더 (A~AX)
OFFICIAL_HEADERS = [
    "item_number",                    # A
    "seller_unique_item_id",          # B
    "category_number",                # C
    "brand_number",                   # D
    "item_name",                      # E
    "item_promotion_name",            # F
    "item_status",                    # G
    "start_date",                     # H
    "end_date",                       # I
    "price_yen",                      # J
    "retail_price_yen",               # K
    "taxrate",                        # L
    "quantity",                       # M
    "option_info",                    # N
    "additional_option_info",         # O
    "additional_option_text",         # P
    "image_main_url",                 # Q
    "image_other_url",                # R
    "video_url",                      # S
    "image_option_info",              # T
    "image_additional_option_info",   # U
    "header_html",                    # V
    "footer_html",                    # W
    "item_description",               # X
    "Shipping_number",                # Y
    "option_number",                  # Z
    "available_shipping_date",        # AA
    "desired_shipping_date",          # AB
    "search_keyword",                 # AC
    "item_condition_type",            # AD
    "origin_type",                    # AE
    "origin_region_id",               # AF
    "origin_country_id",              # AG
    "origin_others",                  # AH
    "medication_type",                # AI
    "item_weight",                    # AJ
    "item_material",                  # AK
    "model_name",                     # AL
    "external_product_type",          # AM
    "external_product_id",            # AN
    "manufacture_date",               # AO
    "expiration_date_type",           # AP
    "expiration_date_MFD",            # AQ
    "expiration_date_PAO",            # AR
    "expiration_date_EXP",            # AS
    "under18s_display",               # AT
    "A/S_info",                       # AU
    "buy_limit_type",                 # AV
    "buy_limit_date",                 # AW
    "buy_limit_qty",                  # AX
]

# ── 카테고리 매핑 ────────────────────────────────────────────
CATEGORY_MAP = {
    "스킨케어": "100000043",
    "토너/스킨": "100000043",
    "세럼/에센스": "100000043",
    "크림": "100000043",
    "로션/에멀전": "100000043",
    "클렌징": "100000043",
    "마스크팩": "100000043",
    "선케어": "100000043",
    "립메이크업": "100000044",
    "아이메이크업": "100000044",
    "페이스메이크업": "100000044",
    "베이스메이크업": "100000044",
    "헤어케어": "100000046",
    "헤어샴푸": "100000046",
    "헤어트리트먼트": "100000046",
    "바디케어": "100000047",
    "바디워시": "100000047",
    "바디로션": "100000047",
    "향수": "100000045",
    "네일": "100000044",
    "뷰티툴": "100000044",
    "세트/키트": "100000043",
}
DEFAULT_CATEGORY = "100000043"

# ── 브랜드 매핑 ──────────────────────────────────────────────
BRAND_MAP = {
    "ANUA": "QBR00001",
    "COSRX": "QBR00002",
    "INNISFREE": "QBR00003",
    "LANEIGE": "QBR00004",
    "MISSHA": "QBR00005",
    "ETUDE": "QBR00006",
    "SOME BY MI": "QBR00007",
    "BEAUTY OF JOSEON": "QBR00008",
    "TORRIDEN": "QBR00009",
    "ISNTREE": "QBR00010",
    "MEDICUBE": "QBR00011",
    "NUMBUZIN": "QBR00012",
    "SKIN1004": "QBR00013",
    "ROUNDLAB": "QBR00014",
    "DR.JART+": "QBR00015",
}

# ── 무게 매핑 (g 단위 내부 관리) ─────────────────────────────
WEIGHT_MAP_G = {
    "100000043": 300,
    "100000044": 200,
    "100000045": 200,
    "100000046": 400,
    "100000047": 350,
}
DEFAULT_WEIGHT_G = 300

# ── 상품명 금지 키워드 (큐텐 정책) ────────────────────────────
FORBIDDEN_PROMO_WORDS = [
    "割引", "特価", "セール", "SALE", "sale", "Sale",
    "激安", "最安", "最安値", "格安",
    "限定", "期間限定",
    "送料無料", "無料配送",
    "ポイント", "クーポン",
    "1+1", "2+1", "おまけ",
    "人気No.1", "ランキング1位", "売れ筋",
]


# ══════════════════════════════════════════════════════════════
# 헬퍼 함수
# ══════════════════════════════════════════════════════════════

def _safe_str(value) -> str:
    """bytes나 비문자열 타입을 안전하게 str로 변환"""
    if value is None:
        return ""
    if isinstance(value, bytes):
        return ""
    if isinstance(value, float) and (value != value):  # NaN 체크
        return ""
    return str(value) if not isinstance(value, str) else value


def _match_qoo10_category(item: dict) -> str:
    """상품의 카테고리를 Qoo10 카테고리 번호로 매핑"""
    category = item.get("category", "")
    return CATEGORY_MAP.get(category, DEFAULT_CATEGORY)


def _match_brand_code(item: dict) -> str:
    """브랜드명을 Qoo10 브랜드 코드로 매핑"""
    brand = item.get("brand", "").upper().strip()
    return BRAND_MAP.get(brand, "")


def _get_item_weight_kg(category_code: str) -> float:
    """
    ★ v0.9.4: 카테고리별 무게를 kg 단위로 반환 (소수점 둘째자리)
    Qoo10 규격: 숫자 2자리 이내, 소수점 둘째자리까지, 최대 30kg
    """
    weight_g = WEIGHT_MAP_G.get(category_code, DEFAULT_WEIGHT_G)
    return round(weight_g / 1000, 2)


def _truncate_item_name(name: str, max_len: int = 50) -> str:
    """상품명에서 홍보 문구 제거 후 길이 제한"""
    if not name:
        return ""
    for word in FORBIDDEN_PROMO_WORDS:
        name = name.replace(word, "")
    name = re.sub(r'\s+', ' ', name).strip()
    if len(name) > max_len:
        name = name[:max_len - 1] + "…"
    return name


def _clean_html(raw_html: str, max_len: int = 2500) -> str:
    """
    ★ v0.9.4: HTML 콘텐츠에서 Python 코드 혼입 제거 + 글자 수 제한
    product_analyzer.py가 header_html에 Python 변수 할당 코드를 포함시키는 버그 대응
    """
    if not raw_html:
        return ""
    if isinstance(raw_html, bytes):
        return ""

    # Python 코드 패턴 제거 (marketing_html = "", if marketing_message: 등)
    python_patterns = [
        r'#\s*마케팅.*$',
        r'#\s*리뷰.*$',
        r'marketing_html\s*=.*$',
        r'social_proof_html\s*=.*$',
        r'social_parts\s*=.*$',
        r'items_str\s*=.*$',
        r'if\s+marketing_message:.*$',
        r'if\s+review_highlight:.*$',
        r'if\s+texture_jp:.*$',
        r'if\s+reviews_summary.*$',
        r'if\s+social_parts:.*$',
        r'for\s+sp\s+in\s+social_parts:.*$',
        r'items_str\s*\+=.*$',
        r'social_parts\.append\(.*$',
        r'marketing_html\s*=\s*\(.*$',
        r'social_proof_html\s*=\s*\(.*$',
        r'\+\s*items_str\s*\+.*$',
    ]

    lines = raw_html.split('\n')
    cleaned_lines = []
    for line in lines:
        stripped = line.strip()
        is_python = False
        for pattern in python_patterns:
            if re.match(pattern, stripped):
                is_python = True
                break
        # 추가: 들여쓰기 + 따옴표로 시작하는 Python 문자열 연결
        if stripped.startswith("'") and stripped.endswith("'"):
            # HTML 태그가 포함되어 있으면 유지, 아니면 제거
            if '<' not in stripped:
                is_python = True
        if not is_python:
            cleaned_lines.append(line)

    result = '\n'.join(cleaned_lines).strip()

    # 글자 수 제한
    if len(result) > max_len:
        # HTML 태그가 중간에 잘리지 않도록 마지막 닫힌 태그 위치에서 자름
        truncated = result[:max_len]
        last_close = truncated.rfind('>')
        if last_close > max_len - 200:
            result = truncated[:last_close + 1]
        else:
            result = truncated

    return result


def _extract_search_keywords(item: dict, category: str = "") -> str:
    """검색 키워드 추출 (30자 제한, 약기법 필터 포함)"""
    keywords = []

    brand = item.get("brand", "")
    if brand:
        keywords.append(brand)

    name_jp = item.get("name_jp", item.get("name", ""))
    if name_jp:
        clean_name = re.sub(r'【.*?】', '', name_jp).strip()
        clean_name = re.sub(r'\d+\s*(?:ml|g|mg|oz|L)\b', '', clean_name, flags=re.IGNORECASE)
        if clean_name:
            keywords.append(clean_name)

    category_kr = item.get("category", "")
    cat_kw_map = {
        "스킨케어": "韓国コスメ スキンケア",
        "토너/스킨": "韓国コスメ 化粧水",
        "세럼/에센스": "韓国コスメ 美容液",
        "크림": "韓国コスメ クリーム",
        "마스크팩": "韓国コスメ パック",
        "선케어": "韓国コスメ 日焼け止め",
        "클렌징": "韓国コスメ クレンジング",
    }
    cat_kw = cat_kw_map.get(category_kr, "韓国コスメ")
    keywords.append(cat_kw)

    search_kw = " ".join(keywords)

    if search_kw and YAKUJIHO_ENABLED:
        search_kw, _, count = sanitize_jp(search_kw)
        if count > 0:
            logger.info(f"[약기법] 검색 키워드에서 {count}건 치환")

    if len(search_kw) > 30:
        search_kw = search_kw[:30]
        last_space = search_kw.rfind(" ")
        if last_space > 20:
            search_kw = search_kw[:last_space]

    return search_kw.strip()


# ══════════════════════════════════════════════════════════════
# 등록코드 관리
# ══════════════════════════════════════════════════════════════

def _load_registered_codes() -> set:
    """기존 Qoo10 등록 상품코드 로드"""
    try:
        with open("registered_codes.json", "r") as f:
            codes = set(json.load(f))
            logger.info(f"[등록코드] {len(codes)}건 로드")
            return codes
    except FileNotFoundError:
        logger.info("[등록코드] registered_codes.json 없음 — 첫 실행으로 간주")
        return set()
    except Exception as e:
        logger.warning(f"[등록코드] 로드 실패: {e} — 필터 스킵")
        return set()


def _save_registered_codes(new_codes: list):
    """신규 등록 코드를 registered_codes.json에 추가"""
    existing = set()
    try:
        with open("registered_codes.json", "r") as f:
            existing = set(json.load(f))
    except FileNotFoundError:
        pass

    updated = sorted(existing | set(new_codes))
    with open("registered_codes.json", "w") as f:
        json.dump(updated, f, indent=2, ensure_ascii=False)
    logger.info(f"[등록코드] {len(new_codes)}건 추가 → 총 {len(updated)}건 저장")


# ══════════════════════════════════════════════════════════════
# 1. Qoo10 엑셀 생성 (v0.9.4)
# ══════════════════════════════════════════════════════════════

def generate_qoo10_excel(items: list) -> BytesIO:
    """Qoo10 J·QSM 대량등록용 엑셀 파일 생성 (공식 50컬럼)"""

    # ── 기존 등록 상품 제외 ──
    registered = _load_registered_codes()
    if registered:
        before = len(items)
        items = [
            i for i in items
            if f"KJ{i.get('item_id', '') or i.get('product_id', '')}" not in registered
        ]
        skipped = before - len(items)
        if skipped:
            logger.info(f"[기존제외] {before}건 → {len(items)}건 (기존 등록 {skipped}건 제외)")

    if not items:
        logger.warning("[결과] 신규 상품 0건 — 엑셀 생성 스킵")
        return None

    # ── openpyxl import ──
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
    except ImportError:
        logger.error("openpyxl 미설치 – pip install openpyxl")
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Qoo10_Upload"

    # ── 헤더 (1행) ──
    header_font = Font(bold=True, size=10)
    for col, header in enumerate(OFFICIAL_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # ── 데이터 행 ──
    cat_stats = {}
    brand_matched = 0
    weight_stats = {}
    yakujiho_stats = 0
    thumb_fallback_count = 0
    html_truncated_count = 0

    for row_idx, item in enumerate(items, 2):
        # 카테고리
        qoo10_cat = _match_qoo10_category(item)
        cat_stats[qoo10_cat] = cat_stats.get(qoo10_cat, 0) + 1

        # 브랜드
        qoo10_brand = _match_brand_code(item)
        if qoo10_brand:
            brand_matched += 1

        # ★ v0.9.4: 무게 (kg 단위)
        weight_kg = _get_item_weight_kg(qoo10_cat)
        weight_key = str(weight_kg)
        weight_stats[weight_key] = weight_stats.get(weight_key, 0) + 1

        # ★ v0.9.4: 이미지 — 로컬 경로면 원본 URL로 폴백
        thumbnail = item.get("thumbnail", "") or item.get("image_url", "")
        thumb_processed = item.get("thumbnail_processed", "")

        if isinstance(thumb_processed, bytes):
            thumb_processed = ""

        if thumb_processed and isinstance(thumb_processed, str) and thumb_processed.startswith("http"):
            main_image = thumb_processed
        elif thumbnail and isinstance(thumbnail, str) and thumbnail.startswith("http"):
            main_image = thumbnail
            if thumb_processed:
                thumb_fallback_count += 1
                logger.debug(f"[이미지] 로컬 경로 → 원본 URL 폴백: {thumb_processed}")
        else:
            main_image = ""

        detail_images = item.get("detail_images", [])
        other_images = "$$".join(
            [img for img in detail_images[:20] if isinstance(img, str) and img.startswith("http")]
        ) if detail_images else ""

        # 옵션
        detail = item.get("detail_info", {})
        opt_name = detail.get("option1_name", "")
        opt_values = detail.get("option1_values", [])
        option_str = ""
        if opt_name and opt_values:
            sell_price_for_opt = item.get("sell_price_jpy", 0)
            option_parts = [f"{v}^{sell_price_for_opt}^99^0^0" for v in opt_values]
            option_str = f"{opt_name}:#{'$$'.join(option_parts)}"

        # 가격
        sell_price = item.get("sell_price_jpy", 0)
        retail_price = item.get("consumer_price_jpy", 0)
        if retail_price <= sell_price:
            retail_price = math.ceil(sell_price * 1.3)

        # 상품명 (홍보문구 제거 + 50자 제한)
        name_jp = item.get("name_jp", item.get("name", ""))
        item_name = _truncate_item_name(name_jp)

        # 약기법 필터 (상품명)
        if item_name and YAKUJIHO_ENABLED:
            item_name, _, name_filter_count = sanitize_jp(item_name)
            if name_filter_count > 0:
                yakujiho_stats += name_filter_count
                logger.info(f"[약기법] 상품명에서 {name_filter_count}건 치환: {item_name[:30]}")

        # 검색어 (약기법 필터 포함)
        search_kw = _extract_search_keywords(item, qoo10_cat)

        # ★ v0.9.4: header_html, footer_html — Python 코드 제거 + 2,500자 제한
        raw_header = item.get("header_html", "")
        raw_footer = item.get("footer_html", "")
        header_html = _clean_html(raw_header, HEADER_MAX_LEN)
        footer_html = _clean_html(raw_footer, FOOTER_MAX_LEN)

        if raw_header and len(raw_header) > HEADER_MAX_LEN:
            html_truncated_count += 1
        if raw_footer and len(raw_footer) > FOOTER_MAX_LEN:
            html_truncated_count += 1

        # seller_unique_item_id
        seller_uid = f"KJ{item.get('item_id', '') or item.get('product_id', '')}"

        # ── 50컬럼 행 데이터 (A~AX) ──
        row_data = [
            '',                                         # A  item_number (신규→공란)
            seller_uid,                                 # B  seller_unique_item_id
            str(qoo10_cat),                             # C  category_number
            str(qoo10_brand),                           # D  brand_number
            item_name,                                  # E  item_name
            '韓国コスメ 正規品',                           # F  item_promotion_name
            'Y',                                        # G  item_status
            '',                                         # H  start_date
            '2030-12-31',                               # I  end_date
            str(sell_price),                            # J  price_yen
            str(retail_price),                          # K  retail_price_yen
            '',                                         # L  taxrate
            '100',                                      # M  quantity
            option_str,                                 # N  option_info
            '',                                         # O  additional_option_info
            '',                                         # P  additional_option_text
            main_image,                                 # Q  image_main_url ★ v0.9.4
            other_images,                               # R  image_other_url
            '',                                         # S  video_url
            '',                                         # T  image_option_info
            '',                                         # U  image_additional_option_info
            header_html,                                # V  header_html ★ v0.9.4
            footer_html,                                # W  footer_html ★ v0.9.4
            _safe_str(item.get('detail_html', '')),     # X  item_description
            QOO10_KSE_SHIPPING_CODE,                    # Y  Shipping_number ★ v0.9.4 (숫자코드)
            '',                                         # Z  option_number
            '3',                                        # AA available_shipping_date
            '7',                                        # AB desired_shipping_date
            search_kw,                                  # AC search_keyword
            '1',                                        # AD item_condition_type
            '2',                                        # AE origin_type
            '',                                         # AF origin_region_id
            'KR',                                       # AG origin_country_id
            '',                                         # AH origin_others
            '',                                         # AI medication_type
            str(weight_kg),                             # AJ item_weight ★ v0.9.4 (kg)
            '',                                         # AK item_material
            '',                                         # AL model_name
            '',                                         # AM external_product_type
            '',                                         # AN external_product_id
            '',                                         # AO manufacture_date
            '',                                         # AP expiration_date_type
            '',                                         # AQ expiration_date_MFD
            '',                                         # AR expiration_date_PAO
            '',                                         # AS expiration_date_EXP
            'N',                                        # AT under18s_display
            '',                                         # AU A/S_info
            '',                                         # AV buy_limit_type
            '',                                         # AW buy_limit_date
            '',                                         # AX buy_limit_qty
        ]

        # ★ v0.9.4: 전체 row bytes 방어
        for col, value in enumerate(row_data, 1):
            if isinstance(value, bytes):
                value = ""
            ws.cell(row=row_idx, column=col, value=value)

    # ── 열 너비 ──
    col_widths = {
        "A": 12, "B": 18, "C": 15, "D": 12, "E": 50,
        "F": 22, "G": 8, "H": 14, "I": 14, "J": 12,
        "K": 12, "L": 8, "M": 8, "N": 40, "Q": 50,
        "R": 50, "V": 30, "W": 30, "X": 30,
        "Y": 12, "AC": 40, "AG": 8, "AJ": 8,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # ── BytesIO로 출력 ──
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # ── 로그 ──
    logger.info(f"[엑셀] Qoo10 업로드 파일 생성: {len(items)}건 (50컬럼)")
    logger.info(f"[엑셀] 브랜드 매칭: {brand_matched}/{len(items)}건")
    if yakujiho_stats > 0:
        logger.info(f"[엑셀] 약기법 필터: 총 {yakujiho_stats}건 키워드 치환")
    if thumb_fallback_count > 0:
        logger.info(f"[엑셀] 이미지 폴백: {thumb_fallback_count}건 (로컬경로→원본URL)")
    if html_truncated_count > 0:
        logger.info(f"[엑셀] HTML 잘림: {html_truncated_count}건 (2500자 초과)")
    top5_cat = sorted(cat_stats.items(), key=lambda x: -x[1])[:5]
    logger.info(f"[엑셀] 카테고리 TOP5: {dict(top5_cat)}")
    logger.info(f"[엑셀] 무게 분포(kg): {dict(sorted(weight_stats.items()))}")

    return output


# ══════════════════════════════════════════════════════════════
# 2. Google Drive 업로드
# ══════════════════════════════════════════════════════════════

def upload_to_drive(excel_bytes: BytesIO, filename: str) -> str:
    """엑셀 파일을 Google Drive에 업로드 (실패 시 빈 문자열 반환)"""
    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaIoBaseUpload
        import json as _json

        sa_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "")
        if not sa_json:
            raise ValueError("GOOGLE_SERVICE_ACCOUNT_JSON 환경변수 없음")

        sa_info = _json.loads(sa_json)
        credentials = service_account.Credentials.from_service_account_info(
            sa_info, scopes=["https://www.googleapis.com/auth/drive.file"]
        )
        service = build("drive", "v3", credentials=credentials)

        folder_id = os.environ.get("GOOGLE_DRIVE_FOLDER_ID", "")
        file_metadata = {
            "name": filename,
            "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }
        if folder_id:
            file_metadata["parents"] = [folder_id]

        media = MediaIoBaseUpload(
            excel_bytes,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        file = (
            service.files()
            .create(body=file_metadata, media_body=media, fields="id,webViewLink")
            .execute()
        )

        return file.get("webViewLink", f"https://drive.google.com/file/d/{file.get('id')}")

    except Exception as e:
        logger.warning(f"[Drive] 업로드 실패: {e}")
        return ""


# ══════════════════════════════════════════════════════════════
# 3. 통합 실행
# ══════════════════════════════════════════════════════════════

def generate_and_upload(items: list) -> dict:
    """최종 통과 상품 → 엑셀 생성 → Drive 업로드 (항상 artifacts에도 저장)"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"qoo10_upload_{timestamp}.xlsx"

    excel_bytes = generate_qoo10_excel(items)
    if excel_bytes is None:
        return {"item_count": 0, "filename": "", "drive_url": "", "success": False}

    # ── 항상 artifacts/ 에 로컬 저장 ──
    os.makedirs("artifacts", exist_ok=True)
    local_path = f"artifacts/{filename}"
    excel_bytes.seek(0)
    with open(local_path, "wb") as f:
        f.write(excel_bytes.read())
    logger.info(f"[저장] artifacts 로컬 저장 완료: {local_path}")

    # ── 업로드한 seller_code를 registered_codes.json에 자동 추가 ──
    new_codes = [
        f"KJ{i.get('item_id', '') or i.get('product_id', '')}"
        for i in items
        if i.get('item_id') or i.get('product_id')
    ]
    if new_codes:
        _save_registered_codes(new_codes)

    # ── Drive 업로드 시도 ──
    excel_bytes.seek(0)
    drive_url = upload_to_drive(excel_bytes, filename)
    is_local = not drive_url or drive_url.startswith("LOCAL:")

    result = {
        "item_count": len(items),
        "filename": filename,
        "drive_url": drive_url if not is_local else "",
        "local_path": local_path,
        "success": True,
        "is_local": is_local,
    }

    if drive_url and not is_local:
        logger.info(f"[업로드] Drive 성공: {len(items)}건 → {drive_url}")
        logger.info(f"[업로드] artifacts 백업: {local_path}")
    else:
        logger.info(f"[업로드] Drive 스킵/실패 → GitHub Actions Artifacts 탭에서 다운로드: {local_path}")

    return result


# ══════════════════════════════════════════════════════════════
# 직접 실행 (테스트)
# ══════════════════════════════════════════════════════════════
if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )

    test_items = [
        {
            "name": "어성초 77% 수딩 토너",
            "name_jp": "【ANUA】ドクダミ77% スージングトナー 250ml",
            "brand": "ANUA",
            "category": "토너/스킨",
            "item_id": "12345",
            "product_id": "12345",
            "supply_price": 8000,
            "sell_price_jpy": 2600,
            "consumer_price_jpy": 0,
            "margin_rate": 0.217,
            "final_score": 82,
            "grade": "A",
            "thumbnail": "https://example.com/thumb1.jpg",
            "thumbnail_processed": "artifacts/thumbnails/thumb_12345.jpg",
            "detail_images": [
                "https://example.com/detail1.jpg",
                "https://example.com/detail2.jpg",
            ],
            "detail_html": "<div>テスト商品説明</div>",
            "header_html": "<div>テスト ヘッダー</div>",
            "footer_html": "<div>テスト フッター</div>",
            "detail_info": {
                "option1_name": "容量",
                "option1_values": ["250ml", "500ml"],
            },
        },
        {
            "name": "안티에이징 화이트닝 크림",
            "name_jp": "【TEST】アンチエイジング ホワイトニング クリーム 50ml",
            "brand": "TEST",
            "category": "크림",
            "item_id": "99999",
            "product_id": "99999",
            "supply_price": 12000,
            "sell_price_jpy": 3800,
            "consumer_price_jpy": 0,
            "margin_rate": 0.25,
            "final_score": 75,
            "grade": "B",
            "thumbnail": "https://example.com/thumb2.jpg",
            "thumbnail_processed": b'\xff\xd8\xff\xe0test_bytes',  # 바이너리 테스트
            "detail_images": [],
            "detail_html": "<div>テスト</div>",
            "header_html": "marketing_html = \"\"\nif marketing_message:\n<div>Real Header</div>",
            "footer_html": "<div>Real Footer</div>",
            "detail_info": {},
        },
    ]

    print(f"=== uploader_qoo10.py v0.9.4 테스트 ===")
    print(f"헤더 수: {len(OFFICIAL_HEADERS)}컬럼 (A~AX)")
    print(f"약기법 필터 활성화: {YAKUJIHO_ENABLED}")
    print(f"배송비 코드: {QOO10_KSE_SHIPPING_CODE}")
    print()

    excel = generate_qoo10_excel(test_items)
    if excel:
        test_filename = "test_qoo10_upload.xlsx"
        with open(test_filename, "wb") as f:
            f.write(excel.read())
        print(f"테스트 엑셀 생성 완료: {test_filename}")
        print(f"생성 건수: {len(test_items)}건")
        print()
        print("검증 포인트:")
        print("  1. Q열: 로컬 경로 대신 https:// URL이 들어갔는지")
        print("  2. V/W열: Python 코드 제거 + 2500자 이내인지")
        print("  3. Y열: 숫자 배송비 코드인지")
        print("  4. AJ열: kg 단위(0.3 등)인지")
        print("  5. bytes 값이 빈 문자열로 처리됐는지")
    else:
        print("엑셀 생성 실패 (신규 상품 0건 또는 openpyxl 미설치)")